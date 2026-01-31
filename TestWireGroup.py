from PyQt5.QtWidgets import (
    QWidget, QGroupBox, QLabel, 
    QGridLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QVBoxLayout, QHBoxLayout, QSpacerItem, 
    QSizePolicy, QLineEdit, QMessageBox, QFileDialog,
    QAbstractItemView, QCheckBox
)
from PyQt5.QtGui import QPixmap, QColor, QIcon, QFont
from PyQt5.QtCore import Qt, pyqtSignal

import csv
import os

from SystemModul import icon
from SystemModul import Constants 

from MessageWindows import WarningWindow
from MessageWindows import DangerWindow
from MessageWindows import SuccessWindow
from MessageWindows import InfoWindow



class TestWireGroup(QWidget):  # QWidget –≤–º–µ—Å—Ç–æ QMainWindow
    """–í–∏–¥–∂–µ—Ç –¥–ª—è —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏—è —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ –ø—Ä–æ–∑–≤–æ–Ω–∫–∏ –ø—Ä–æ–≤–æ–¥–æ–≤"""

    # –°–∏–≥–Ω–∞–ª—ã –¥–ª—è —Å–≤—è–∑–∏ —Å –¥—Ä—É–≥–∏–º–∏ –∫–æ–º–ø–æ–Ω–µ–Ω—Ç–∞–º–∏
    
    def __init__(self, parent=None):
        super().__init__(parent)

        self.const = Constants()
        self.icon = icon()

        self.min_size_x = 30 
        self.min_size_y = 30

        self.read_bit_rows = []
        self.wire_data_from_file = []

        # 
        self.update_data_to_test = 0
        self.update_data_to_test_text = ""

        self.accord_data = []


        self.init_ui()

        # —Ç—É—Ç –æ–±–Ω–æ–≤–∏—Ç—å —Å—Ç–∞—Ç—É—Å
        # self.to_update_data_to_test()

    
    def init_ui(self):
        # –û—Å–Ω–æ–≤–Ω–æ–π layout
        main_layout = QVBoxLayout(self)
        
        wires_group = QGroupBox(self.const.TEST_TABLE_GROUP_TITLE)
        wires_layout = QVBoxLayout()
        
        # –¢–∞–±–ª–∏—Ü–∞ –¥–ª—è –æ—Ç–æ–±—Ä–∞–∂–µ–Ω–∏—è –ø—Ä–æ–≤–æ–¥–æ–≤
        self.wires_table = QTableWidget()
        self.wires_table.setColumnCount(3)
        self.wires_table.setHorizontalHeaderLabels(
            self.const.TABLE_HEADERS["wire_test"]
        )
        header = self.wires_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignLeft)
        
        self.wires_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.wires_table.setSelectionBehavior(QTableWidget.SelectRows)


        # –ö–Ω–æ–ø–∫–∏ —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è
        buttons_group_main = QGroupBox()
        buttons_group_main.setMaximumSize(1000, 220)
        buttons_layout_main = QGridLayout(buttons_group_main)

        # 1
        buttons_group_1 = QGroupBox()
        buttons_layout_1 = QGridLayout(buttons_group_1)

        self.line_test_file = QLineEdit()
        self.line_test_file.setStyleSheet('background : #ccc; ')
        self.line_test_file.setReadOnly(1)
        self.line_test_file.setPlaceholderText(self.const.PLACEHOLDER_TEXTS["test_file_select"])

        self.open_button = QPushButton(self.const.BUTTON_TEXTS["open"])
        self.open_button.setIcon(self.icon.open_folder_icon)
        self.open_button.clicked.connect(self.read_from_csv)
        
        self.check_button = QPushButton(self.const.BUTTON_TEXTS["check"])
        self.check_button.setIcon(self.icon.search_icon)
        self.check_button.clicked.connect(self.do_check)

        self.save_button = QPushButton(self.const.BUTTON_TEXTS["save_check_result"])
        self.save_button.setIcon(self.icon.save_icon)
        self.save_button.clicked.connect(self.save_check_result)

        buttons_layout_1.addWidget(self.line_test_file,  1, 0, 1, 1)
        buttons_layout_1.addWidget(self.open_button,     1, 1, 1, 1)
        buttons_layout_1.addWidget(self.check_button,    2, 0, 1, 2)
        buttons_layout_1.addWidget(self.save_button,     3, 0, 1, 2)
        
        # 2
        check_box_group = QGroupBox()
        check_box_layout = QGridLayout(check_box_group)

        self.check_box_num =  QCheckBox(self.const.BUTTON_TEXTS["num_pin"], self)
        self.check_box_num.toggle()
        self.check_box_num.stateChanged.connect(self.update_buttons_text)

        self.check_box_name = QCheckBox(self.const.BUTTON_TEXTS["socket_name"], self)
        self.check_box_name.toggle()
        self.check_box_name.stateChanged.connect(self.update_buttons_text)

        check_box_layout.addWidget(self.check_box_num,  0, 0, 1, 1)
        check_box_layout.addWidget(self.check_box_name,  0, 1, 1, 1)


        buttons_layout_main.addWidget(buttons_group_1)
        buttons_layout_main.addWidget(check_box_group)

        spacerItem = QSpacerItem(20, 40, QSizePolicy.Maximum, QSizePolicy.Expanding)
        buttons_layout_main.addItem(spacerItem)

        wires_layout.addWidget(self.wires_table)
        wires_layout.addWidget(buttons_group_main)
        wires_group.setLayout(wires_layout)

        main_layout.addWidget(wires_group)


    # —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Å–∏–≥–Ω–∞–ª–∞
    def process_accord_data(self, accord_data):
        self.accord_data = accord_data

# —Ç—É—Ç –æ–Ω–∞ —Å–∫–æ—Ä–µ–π –≤—Å–µ–≥–æ –Ω–µ –ø–æ—Ç—Ä–µ–±—É–µ—Ç—Å—è 
    # # —Ç–∞–∫–∞—è –∂–µ —Ñ—É–Ω–∫—Ü–∏—è, –∫–∞–∫ –∏ –≤ read
    # def fill_table_from_accord_data(self):
    #     """–ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ —Ç–∞–±–ª–∏—Ü—ã –¥–∞–Ω–Ω—ã–º–∏ –∏–∑ —Ñ–∞–π–ª–∞ —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤–∏–π (—Ç–æ–ª—å–∫–æ 2 —Å—Ç–æ–ª–±—Ü–∞)"""
    #     # –û—á–∏—â–∞–µ–º —Ç–∞–±–ª–∏—Ü—É
    #     self.wires_table.setRowCount(0)
        
    #     if not self.accord_data:
    #         return
        
    #     # –ï—Å–ª–∏ –ø–µ—Ä–≤–∞—è —Å—Ç—Ä–æ–∫–∞ - –∑–∞–≥–æ–ª–æ–≤–∫–∏
    #     has_headers = True
    #     # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —è–≤–ª—è–µ—Ç—Å—è –ª–∏ –ø–µ—Ä–≤–∞—è —Å—Ç—Ä–æ–∫–∞ –∑–∞–≥–æ–ª–æ–≤–∫–∞–º–∏ (—Å–æ–¥–µ—Ä–∂–∏—Ç —Ç–µ–∫—Å—Ç, –∞ –Ω–µ —á–∏—Å–ª–∞)
    #     if self.accord_data and len(self.accord_data[0]) > 0:
    #         first_cell = self.accord_data[0][0]
    #         if first_cell and first_cell.lower() in ["—Ä–∞–∑—ä–µ–º", "socket", "connector", "–≤—ã–≤–æ–¥", "pin"]:
    #             has_headers = True
    #             headers = self.accord_data[0]
    #             data_rows = self.accord_data[1:]
    #         else:
    #             has_headers = False
    #             data_rows = self.accord_data
    #     else:
    #         has_headers = False
    #         data_rows = self.accord_data
        
    #     # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Å—Ç—Ä–æ–∫
    #     self.wires_table.setRowCount(len(data_rows))
        
    #     # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏ —Ç–∞–±–ª–∏—Ü—ã
    #     if has_headers and len(headers) >= 2:
    #         # self.wires_table.setHorizontalHeaderLabels([headers[0], headers[1], "–í—ã–≤–æ–¥—ã"])
    #         self.wires_table.setHorizontalHeaderLabels([
    #             "–†–∞–∑—ä–µ–º", "–í—ã–≤–æ–¥", "–í—ã–≤–æ–¥"
    #         ])
        
    #     header = self.wires_table.horizontalHeader()
    #     header.setDefaultAlignment(Qt.AlignLeft)

    #     # –ó–∞–ø–æ–ª–Ω—è–µ–º –¥–∞–Ω–Ω—ã–µ
    #     for row_idx, row_data in enumerate(data_rows):
    #         # –†–∞–∑—ä–µ–º (—Å—Ç–æ–ª–±–µ—Ü 0)
    #         if len(row_data) > 0:
    #             item_socket = QTableWidgetItem(row_data[0])
    #             item_socket.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
    #             self.wires_table.setItem(row_idx, 0, item_socket)
            
    #         # –í—ã–≤–æ–¥ (—Å—Ç–æ–ª–±–µ—Ü 1)
    #         if len(row_data) > 1:
    #             item_pin = QTableWidgetItem(row_data[1])
    #             item_pin.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
    #             self.wires_table.setItem(row_idx, 1, item_pin)
            
    #     # –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –ø–æ–¥–≥–æ–Ω—è–µ–º —à–∏—Ä–∏–Ω—É —Å—Ç–æ–ª–±—Ü–æ–≤
    #     self.wires_table.resizeColumnsToContents()
        
    #     # –î–µ–ª–∞–µ–º —Ç—Ä–µ—Ç–∏–π —Å—Ç–æ–ª–±–µ—Ü —à–∏—Ä–µ
    #     header = self.wires_table.horizontalHeader()
    #     header.setStretchLastSection(True)




    def read_from_csv(self):
        start_dir = self.const.DIRECTORIES["wire_dataset"]

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "–û—Ç–∫—Ä—ã—Ç—å —Ñ–∞–π–ª",
            start_dir,
            "CSV —Ñ–∞–π–ª—ã (*.csv);;–í—Å–µ —Ñ–∞–π–ª—ã (*.*)"
        )

        if not file_path:
            return  # –û—Ç–º–µ–Ω–∞

        try:
            with open(file_path, "r", encoding="utf-8-sig", newline="") as file:
                reader = csv.reader(file, delimiter=";")
                rows = list(reader)

            if not rows:
                self.WarningWindow  = WarningWindow("–û—à–∏–±–∫–∞. –§–∞–π–ª –ø—É—Å—Ç")
                self.WarningWindow.Window.show()
                return

            headers = rows[0]
            data_rows = rows[1:]

            self.wire_data_from_file = data_rows
            print(self.wire_data_from_file)

            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Å—Ç–æ–ª–±—Ü–æ–≤
            if len(headers) != self.wires_table.columnCount():
                self.DangerWindow = DangerWindow("–û—à–∏–±–∫–∞. –ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç —Ñ–∞–π–ª–∞ (–Ω–µ —Å–æ–≤–ø–∞–¥–∞–µ—Ç –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Å—Ç–æ–ª–±—Ü–æ–≤)")
                self.DangerWindow.Window.show()
                return

            # –û—á–∏—â–∞–µ–º —Ç–∞–±–ª–∏—Ü—É
            self.wires_table.clearContents()
            self.wires_table.setRowCount(0)

            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –≤ —Ç–∞–±–ª–∏—Ü—É
            self.line_test_file.setText(os.path.basename(file_path))

            self.InfoWindow = InfoWindow(f"–§–∞–π–ª {os.path.basename(file_path)} —É—Å–ø–µ—à–Ω–æ –∑–∞–≥—Ä—É–∂–µ–Ω")
            self.InfoWindow.Window.show()

        except Exception as e:
            # QMessageBox.critical(self, "–û—à–∏–±–∫–∞", str(e))
            print(str(e))


    def make_btn_text(self, pin_number, socket_name):
        if self.check_box_num.isChecked() and self.check_box_name.isChecked():
            return f"{pin_number}: {socket_name}"
        elif self.check_box_num.isChecked():
            return f"{pin_number}"
        elif self.check_box_name.isChecked():
            return f"{socket_name}"
        else:
            return ""


    def update_buttons_text(self):
        table = self.wires_table

        for row in range(table.rowCount()):
            cell_widget = table.cellWidget(row, 2)
            if not cell_widget:
                continue

            layout = cell_widget.layout()
            for i in range(layout.count()):
                btn = layout.itemAt(i).widget()
                if not btn:
                    continue

                pin = btn.property("pin_number")
                socket = btn.property("socket_name")

                btn.setText(self.make_btn_text(pin, socket))
                btn.adjustSize()

        table.resizeColumnToContents(2)

            

    def do_check(self):

        total_ok = 0
        total_warning = 0
        total_error = 0

        table = self.wires_table
        table.clearContents()

        # ---------- —Ñ–æ—Ä–º–∏—Ä—É–µ–º —Ñ–∞–∫—Ç–∏—á–µ—Å–∫–∏–µ –∑–∞–º—ã–∫–∞–Ω–∏—è ----------
        intersections_array = []
        for row_index, row in enumerate(self.read_bit_rows):
            mirrored = row[::-1]
            zero_indexes = [i for i, bit in enumerate(mirrored) if bit == 0]
            intersections = [i for i in zero_indexes if i != row_index]
            intersections_array.append(intersections)

        table.setRowCount(len(intersections_array))

        BTN_SIZE = 28
        SPACING = 4
        MARGINS = 8

        btn_color_success = "28A745"
        btn_color_warning = "FFC107"
        btn_color_danger  = "DC3545"

        max_column_width = 0  # –∑–∞–ø–æ–º–Ω–∏–º –º–∞–∫—Å–∏–º–∞–ª—å–Ω—É—é —à–∏—Ä–∏–Ω—É –¥–ª—è —Å—Ç–æ–ª–±—Ü–∞

        # ---------- –æ–±—Ä–∞–±–æ—Ç–∫–∞ —Å—Ç—Ä–æ–∫ ----------
        for i, intersections in enumerate(intersections_array):

            # item_socket = QTableWidgetItem(str(self.wire_data_from_file[i][0]))
            item_socket = QTableWidgetItem(str(self.accord_data[i+1][0]))
            item_socket.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            table.setItem(i, 0, item_socket)

            pin = i + 1

            # --- –Ω–æ–º–µ—Ä –≤—ã–≤–æ–¥–∞ ---
            item_pin = QTableWidgetItem(str(pin))
            item_pin.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            table.setItem(i, 1, item_pin)

            # ---------- –§–ê–ö–¢ ----------
            fact = {j + 1 for j in intersections}

            # ---------- –û–ñ–ò–î–ê–ï–ú–û–ï (–∏–∑ CSV) ----------
            expected = set()
            if i < len(self.wire_data_from_file):
                text = self.wire_data_from_file[i][2]
                for part in text.split(","):
                    part = part.strip()
                    if part.isdigit():
                        expected.add(int(part))

            # ---------- –∞–Ω–∞–ª–∏–∑ ----------
            ok = fact & expected
            warning = expected - fact
            error = fact - expected

            total_ok += len(ok)
            total_warning += len(warning)
            total_error += len(error)

            # ---------- –≤–∏–¥–∂–µ—Ç —Å –∫–Ω–æ–ø–∫–∞–º–∏ ----------
            cell_widget = QWidget()
            layout = QHBoxLayout(cell_widget)
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setSpacing(SPACING)
            # layout.setAlignment(Qt.AlignCenter)
            layout.setAlignment(Qt.AlignLeft)

            all_pins = sorted(fact | expected)
            btn_count = len(all_pins)

            for other_pin in all_pins:

                if other_pin in fact and other_pin in expected:
                    color = btn_color_success     # üü¢ –µ—Å—Ç—å –∏ –æ–∂–∏–¥–∞–ª–∏
                elif other_pin in fact and other_pin not in expected:
                    color = btn_color_danger      # üî¥ –µ—Å—Ç—å, –Ω–æ –Ω–µ –æ–∂–∏–¥–∞–ª–∏
                else:
                    color = btn_color_warning     # üü° –æ–∂–∏–¥–∞–ª–∏, –Ω–æ –Ω–µ—Ç

                # –≤–æ—Ç –∑–¥–µ—Å—å –º–µ–Ω—è—Ç—å —Ç–µ–∫—Å—Ç –≤ –∑–∞–≤–∏—Å–∏–º–æ—Å—Ç–∏ –æ—Ç check_box
                # soket_pin_name = self.wire_data_from_file[other_pin - 1][0] # –ë—ã–ª–æ –∏–∑ —Ñ–∞–π–ª–∞
                soket_pin_name = self.accord_data[other_pin][0] # –±—É–¥–µ—Ç –∏–∑ —Ç–∞–±–ª–∏—Ü—ã —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤–∏–π
                btn_text = self.make_btn_text(other_pin, soket_pin_name)

                btn = QPushButton(btn_text)
                btn.setEnabled(False)

                btn.setProperty("pin_number", other_pin)
                btn.setProperty("socket_name", soket_pin_name)

                btn.setStyleSheet(
                    f"""
                    background-color: #{color};
                    border-radius: 12px;
                    color: white;
                    padding: 6px 12px;
                    """
                )

                btn.adjustSize()

                layout.addWidget(btn)

            table.setCellWidget(i, 2, cell_widget)


        # ---------- –ø—Ä–∏–º–µ–Ω—è–µ–º —à–∏—Ä–∏–Ω—É ----------
        table.resizeColumnToContents(2)

        self.const.ok = total_ok
        self.const.warning = total_warning
        self.const.error = total_error

        # ---------- –∏—Ç–æ–≥–æ–≤–∞—è –æ—Ü–µ–Ω–∫–∞ ----------
        if total_error > 0:
            self.DangerWindow = DangerWindow(
                self.const.MESSAGES["error_check_result"].format(
                    ok=total_ok,
                    warning=total_warning,
                    error=total_error
                )
            )
            self.DangerWindow.Window.show()

        elif total_warning > 0:
            self.WarningWindow = WarningWindow(
                self.const.MESSAGES["warning_check_result"].format(
                    ok=total_ok,
                    warning=total_warning,
                    error=total_error
                )
            )
            self.WarningWindow.Window.show()

        else:
            self.SuccessWindow = SuccessWindow(
                self.const.MESSAGES["success_check_result"].format(
                    ok=total_ok,
                    warning=total_warning,
                    error=total_error
                )
            )
            self.SuccessWindow.Window.show()




    def save_check_result(self):
        if not self.wire_data_from_file or self.wires_table.rowCount() == 0:
            self.WarningWindow = WarningWindow("–ù–µ—Ç –¥–∞–Ω–Ω—ã—Ö –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è")
            self.WarningWindow.Window.show()
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "–°–æ—Ö—Ä–∞–Ω–∏—Ç—å —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã –ø—Ä–æ–≤–µ—Ä–∫–∏",
            "check_result.xlsx",
            "Excel —Ñ–∞–π–ª—ã (*.xlsx)"
        )

        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "–ü—Ä–æ–≤–µ—Ä–∫–∞ –ø—Ä–æ–≤–æ–¥–æ–≤"

            # ---------- –°—Ç–∏–ª–∏ ----------
            header_font = Font(bold=True)
            center_alignment = Alignment(horizontal='center', vertical='center')
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # –¶–≤–µ—Ç–∞ –¥–ª—è –∑–∞–ª–∏–≤–∫–∏ —è—á–µ–µ–∫
            ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # –ó–µ–ª–µ–Ω—ã–π
            warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # –ñ–µ–ª—Ç—ã–π
            error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # –ö—Ä–∞—Å–Ω—ã–π

            # ---------- –ó–∞–≥–æ–ª–æ–≤–∫–∏ ----------
            # headers = ["‚Ññ", "–†–∞–∑—ä—ë–º", "–í—ã–≤–æ–¥", "OK", "WARNING", "ERROR"]
            headers = self.const.EXCEL_HEADERS
            ws.append(headers)
            
            # –°—Ç–∏–ª–∏–∑–∞—Ü–∏—è –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            # ---------- –î–∞–Ω–Ω—ã–µ ----------
            row_num = 2
            total_ok_count = 0
            total_warning_count = 0
            total_error_count = 0
            
            for table_row in range(self.wires_table.rowCount()):
                pin = table_row + 1
                
                # –ü–æ–ª—É—á–∞–µ–º –Ω–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ —Ä–∞–∑—ä–µ–º–∞ –∏–∑ —Ç–∞–±–ª–∏—Ü—ã
                socket_item = self.wires_table.item(table_row, 0)
                socket_name = socket_item.text() if socket_item else ""
                
                # ---------- –û–ñ–ò–î–ê–ï–ú–û–ï ----------
                expected_text = self.wire_data_from_file[table_row][2] if table_row < len(self.wire_data_from_file) else ""
                expected = set()
                for part in expected_text.split(","):
                    part = part.strip()
                    if part.isdigit():
                        expected.add(int(part))

                # ---------- –§–ê–ö–¢–ò–ß–ï–°–ö–û–ï ----------
                fact = set()
                ok_details = []
                warning_details = []
                error_details = []
                
                cell_widget = self.wires_table.cellWidget(table_row, 2)
                if cell_widget:
                    for i in range(cell_widget.layout().count()):
                        btn = cell_widget.layout().itemAt(i).widget()
                        if btn:
                            pin_num = btn.property("pin_number")
                            socket_name_other = btn.property("socket_name") if btn.property("socket_name") else ""
                            fact.add(pin_num)
                            
                            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ç–∏–ø —Å–æ–µ–¥–∏–Ω–µ–Ω–∏—è –ø–æ —Ü–≤–µ—Ç—É –∫–Ω–æ–ø–∫–∏
                            btn_style = btn.styleSheet()
                            
                            # –§–æ—Ä–º–∏—Ä—É–µ–º —Å—Ç—Ä–æ–∫—É —Å –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–µ–π –æ —Å–æ–µ–¥–∏–Ω–µ–Ω–∏–∏
                            if self.check_box_num.isChecked() and self.check_box_name.isChecked():
                                connection_str = f"{pin_num}: {socket_name_other}"
                            elif self.check_box_num.isChecked():
                                connection_str = f"{pin_num}"
                            elif self.check_box_name.isChecked():
                                connection_str = f"{socket_name_other}"
                            else:
                                connection_str = f"{pin_num}"
                            
                            # –†–∞—Å–ø—Ä–µ–¥–µ–ª—è–µ–º –ø–æ —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤—É—é—â–∏–º —Å—Ç–æ–ª–±—Ü–∞–º
                            if "#28A745" in btn_style:  # OK
                                ok_details.append(connection_str)
                                total_ok_count += 1
                            elif "#DC3545" in btn_style:  # ERROR
                                error_details.append(connection_str)
                                total_error_count += 1
                            else:  # WARNING
                                warning_details.append(connection_str)
                                total_warning_count += 1

                # –¢–∞–∫–∂–µ –¥–æ–±–∞–≤–ª—è–µ–º WARNING –¥–ª—è –æ–∂–∏–¥–∞–µ–º—ã—Ö, –Ω–æ –Ω–µ –Ω–∞–π–¥–µ–Ω–Ω—ã—Ö —Å–æ–µ–¥–∏–Ω–µ–Ω–∏–π
                for pin_num in expected:
                    if pin_num not in fact:  # –û–∂–∏–¥–∞–ª–æ—Å—å, –Ω–æ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ
                        if pin_num <= len(self.wire_data_from_file):
                            socket_name_other = self.wire_data_from_file[pin_num - 1][0]
                        else:
                            socket_name_other = f"–ù–µ–∏–∑–≤–µ—Å—Ç–Ω—ã–π ({pin_num})"
                        
                        if self.check_box_num.isChecked() and self.check_box_name.isChecked():
                            warning_details.append(f"{pin_num}: {socket_name_other}")
                        elif self.check_box_num.isChecked():
                            warning_details.append(str(pin_num))
                        elif self.check_box_name.isChecked():
                            warning_details.append(socket_name_other)
                        else:
                            warning_details.append(str(pin_num))
                        
                        # –ù–µ —É–≤–µ–ª–∏—á–∏–≤–∞–µ–º —Å—á–µ—Ç—á–∏–∫ –∑–¥–µ—Å—å, —Ç–∞–∫ –∫–∞–∫ —ç—Ç–æ –Ω–µ —Ñ–∞–∫—Ç–∏—á–µ—Å–∫–æ–µ —Å–æ–µ–¥–∏–Ω–µ–Ω–∏–µ

                # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç—Ä–æ–∫—É –≤ Excel
                ws.append([
                    table_row + 1,  # –ü–æ—Ä—è–¥–∫–æ–≤—ã–π –Ω–æ–º–µ—Ä
                    socket_name,  # –ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ —Ä–∞–∑—ä–µ–º–∞
                    pin,  # –ù–æ–º–µ—Ä –≤—ã–≤–æ–¥–∞
                    "\n".join(ok_details),  # OK —Å–æ–µ–¥–∏–Ω–µ–Ω–∏—è
                    "\n".join(warning_details),  # WARNING —Å–æ–µ–¥–∏–Ω–µ–Ω–∏—è
                    "\n".join(error_details)  # ERROR —Å–æ–µ–¥–∏–Ω–µ–Ω–∏—è
                ])
                
                # –ü—Ä–∏–º–µ–Ω—è–µ–º —Å—Ç–∏–ª–∏ –∫ —è—á–µ–π–∫–∞–º
                for col in range(1, 7):  # 6 —Å—Ç–æ–ª–±—Ü–æ–≤
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # –ó–∞–ª–∏–≤–∫–∞ —è—á–µ–µ–∫ –≤ –∑–∞–≤–∏—Å–∏–º–æ—Å—Ç–∏ –æ—Ç –Ω–∞–ª–∏—á–∏—è –¥–∞–Ω–Ω—ã—Ö
                if error_details:
                    ws.cell(row=row_num, column=6).fill = error_fill  # ERROR
                if warning_details:
                    ws.cell(row=row_num, column=5).fill = warning_fill  # WARNING
                if ok_details and not error_details and not warning_details:
                    ws.cell(row=row_num, column=4).fill = ok_fill  # OK
                
                row_num += 1

            # ---------- –ê–≤—Ç–æ—à–∏—Ä–∏–Ω–∞ —Å —É—á–µ—Ç–æ–º –ø–µ—Ä–µ–Ω–æ—Å–∞ —Ç–µ–∫—Å—Ç–∞ ----------
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                
                for cell in column_cells:
                    try:
                        if cell.value:
                            # –£—á–∏—Ç—ã–≤–∞–µ–º –ø–µ—Ä–µ–Ω–æ—Å —Å—Ç—Ä–æ–∫
                            lines = str(cell.value).split('\n')
                            max_line_length = max(len(line) for line in lines)
                            if max_line_length > max_length:
                                max_length = max_line_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 4, 50)  # –ú–∞–∫—Å–∏–º–∞–ª—å–Ω–∞—è —à–∏—Ä–∏–Ω–∞ 50 —Å–∏–º–≤–æ–ª–æ–≤
                ws.column_dimensions[column].width = adjusted_width

            # ---------- –ó–∞–º–µ—Ä–∑—à–∏–µ –æ–±–ª–∞—Å—Ç–∏ –¥–ª—è —É–¥–æ–±—Å—Ç–≤–∞ –ø—Ä–æ—Å–º–æ—Ç—Ä–∞ ----------
            ws.freeze_panes = "B2"

            # ---------- –ò—Ç–æ–≥–æ–≤–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ ----------
            if row_num > 2:  # –ï—Å–ª–∏ –µ—Å—Ç—å –¥–∞–Ω–Ω—ã–µ
                ws.append([])  # –ü—É—Å—Ç–∞—è —Å—Ç—Ä–æ–∫–∞
                
                stats_row = row_num + 1
                ws.append([
                    "", 
                    "–ò–¢–û–ì–û:", 
                    "", 
                    f"OK: {total_ok_count}", 
                    f"WARNING: {total_warning_count}", 
                    f"ERROR: {total_error_count}"
                ])
                
                # –°—Ç–∏–ª–∏–∑–∞—Ü–∏—è —Å—Ç—Ä–æ–∫–∏ —Å –∏—Ç–æ–≥–∞–º–∏
                for col in range(2, 7):
                    cell = ws.cell(row=stats_row, column=col)
                    cell.font = Font(bold=True)
                    cell.border = border

            wb.save(file_path)

            self.SuccessWindow = SuccessWindow(f"–†–µ–∑—É–ª—å—Ç–∞—Ç—ã –ø—Ä–æ–≤–µ—Ä–∫–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã –≤:\n{os.path.basename(file_path)}")
            self.SuccessWindow.Window.show()

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏: {error_details}")
            self.DangerWindow = DangerWindow(f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è:\n{str(e)}")
            self.DangerWindow.Window.show()